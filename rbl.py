from selenium import webdriver
from selenium.webdriver.common.proxy import Proxy, ProxyType
import time
from selenium.webdriver.common.action_chains import ActionChains 
from selenium.webdriver.common.keys import Keys 
from xlrd import open_workbook
from xlutils.copy import copy
import xlwt
import xlsxwriter



handle=''
title=''
html=''
vendor='RB'
typeOfP =''
tags = ''
published = 'TRUE'
option1_name = ''
option1_value = []
option2_name = ''
option2_value= []
option3_name=''
option3_value= []
variant_sku=''#Add Nothing
variant_grams=''#Add Nothing
vi1='' #Add Nothing
vi2='' 
vi3='deny'
variant_full='manual'
variant_price =''
variant_comapareAt=''
vrs='TRUE'#Add Nothing
vt='TRUE'#Add Nothing
vb=''#Add Nothing
img_src=[]
img_pos='1'#Add Nothing
img_alt=''#Add Nothing
gift_card=''#Add Nothing
seo_title=''#Add Nothing
seo_desc=''#Add Nothing
google_shopping_cat=''#Add Nothing
google_shopping_gender=''#Add Nothing
gs_pin=''#Add Nothing
gs=''#Add Nothing
gs1=''#Add Nothing
gsc='new'#Add Nothing
gscus='FALSE'#Add Nothing
gs91=''#Add Nothing
gs92=''#Add Nothing
gs93=''#Add Nothing
gs94=''#Add Nothing
gs95=''#Add Nothing
gs96=''#Add Nothing
weight_unit ='g'#Add Nothing
tax=''#Add Nothing
cos_item=''#Add Nothing
status='draft'#Add Nothing


tagCount=1


from openpyxl import load_workbook
wb = load_workbook("DataShop.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['A'] 
columnImages = ws['Y']
prev = ''
for x in range(len(column)):

	if column[x].value ==prev:
		ws.write(A)
	prev = column[x].value 
	counter = 1



	# from pydrive.drive import GoogleDrive 
	# from pydrive.auth import GoogleAuth 
	   
	# # For using listdir()  
	# # Below code does the authentication 
	# # part of the code 
	# gauth = GoogleAuth() 
	# # Creates local webserver and auto 
	# # handles authentication. 
	# gauth.LocalWebserverAuth()        
	# drive = GoogleDrive(gauth) 
	   
	# # replace the value of this variable 
	# # with the absolute path of the directory 
	# path = r"C:\Users\Sachin\Desktop"
	# f = drive.CreateFile({'title': handle})
	# f.SetContentFile(img_src) 
	# f.upload()
	# print(f)
	 
	# import base64
	# import json
	# import requests
	# from base64 import b64encode

	# client_id = 'e798d5142339e66'
	# headers = {"Authorization": "Client-ID my-client-id"}
	# api_key = '99870554a23b003dbc297370bbf02241a73e007f'
	# url = "https://api.imgur.com/3/upload"
	# j1 = requests.post(
	# 	url, 
	# 	headers = headers,
	# 	data = {
 #        'key': api_key, 
 #        'image':img_src,
 #        'type': 'url',
 #        'name': handle +'.jpg',
 #        'title': title
 #    }
	# 	)
	# jsonResponse = j1.json()
	# img_src = jsonResponse["data"]["link"]

	#idd = 'VmVeIhy'
	#uurl = https://api.imgur.com/3/image/{{imageHash}}
	#response = requests.request("GET", url, headers=headers, data = payload, files = files)



#except: 
		  
		# Raise UploadError if file is not uploaded. 
	#	raise UploadError("Can't Upload File.") 
    # Due to a known bug in pydrive if we  
    # don't empty the variable used to 
    # upload the files to Google Drive the 
    # file stays open in memory and causes a 
    # memory leak, therefore preventing its  
    # deletion 
	#f = None

	
